from audioop import reverse
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import get_object_or_404, render, redirect
from django.views import View
from .models import Aviso
from .models import Estudiante
from .models import *
from django.db import IntegrityError
from django.views import View
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.staticfiles import finders
from django.urls import reverse
from rest_framework import serializers, viewsets
from .models import Usuarios
import jwt
import datetime
from django.conf import settings
from .models import Canalizacion, BajaAlumnos, AccionTutorial,Usuarios,AtencionIndividual,Periodo,CierreTutorias
from django.views import View
from django.views.decorators.http import require_http_methods
from django.db.models import Count, F, Value
from django.utils.timezone import make_aware
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import cm, inch, mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from reportlab.platypus import Image as ImageCanalizacion
from reportlab.lib.styles import getSampleStyleSheet
from django.utils import timezone


from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from io import BytesIO
import os
from .models import Periodo, AccionTutorial, AtencionIndividual, Usuarios, EvaluacionTutor, Canalizacion, CierreTutorias
from django.contrib import messages #Importamos para presentar mensajes
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator



class TokenSerializer(serializers.ModelSerializer):
    class Meta:
        model = Usuarios
        fields = ['token']

class TokenViewSet(viewsets.ModelViewSet):
    queryset = Usuarios.objects.all()
    serializer_class = TokenSerializer

# Create your views here.
class inicio(View):
   def get(self, request):
        return render(request, 'login.html')

class loginTest(View):
   def get(self, request):
      return render(request,'test/prueba.html')

class aviso(View):
   def get(self, request):
      return render(request,'entrevistas/alumno/aviso.html')

class formulario(View):
   def get(self, request):
      return render(request,'entrevistas/alumno/formulario.html')

class resumenresp(View):
   def get(self, request):
      return render(request,'entrevistas/alumno/resumenresp.html')

class index(View):
   def get(self, request):
      avisos = Aviso.objects.all()
      return render(request,'entrevistas/administrador/index.html', {'avisos': avisos})

class registro(View):
   def get(self, request):
      return render(request,'entrevistas/administrador/registro.html')

class informe(View):
   def get(self, request):
      return render(request,'entrevistas/administrador/informe.html')

class ver(View):
   def get(self, request):
      return render(request,'entrevistas/administrador/ver.html')

#ENTREVISTAS
#ADMINISTRADOR
def registrarAviso(request):
   titulo=request.POST['titulo']
   descripcion=request.POST['descripcion']
   fechaInicio=request.POST['fechaInicio']
   fechaFin=request.POST['fechaFin']

   aviso = Aviso.objects.create(titulo=titulo, descripcion=descripcion, fechaInicio=fechaInicio,fechaFin=fechaFin)
   return redirect('index')

def edicionAviso(request, idAvisos):
    aviso = Aviso.objects.get(idAvisos=idAvisos)
    return render(request, "entrevistas/administrador/editar.html", {"aviso": aviso})

def editarAviso(request):
   idAvisos=request.POST['idAvisos']
   titulo = request.POST['titulo']
   descripcion=request.POST['descripcion']
   fechaInicio=request.POST['fechaInicio']
   fechaFin=request.POST['fechaFin']

   aviso = Aviso.objects.get(idAvisos = idAvisos)
   aviso.titulo = titulo
   aviso.descripcion = descripcion
   aviso.fechaInicio = fechaInicio
   aviso.fechaFin = fechaFin
   aviso.save()
   return redirect('index')

def eliminarAviso(request, idAvisos):
    aviso = Aviso.objects.get(idAvisos=idAvisos)
    aviso.delete()
    return redirect('index')

#ALUMNO
@method_decorator(login_required, name='dispatch')
class aviso(View):
    def get(self, request):
        user = request.user
        try:
            usuario = Usuarios.objects.get(User=user)
            no_control = usuario.noControl
            estudiante_exists = Estudiante.objects.filter(noControl=no_control).exists()
        except Usuarios.DoesNotExist:
            return render(request, 'error.html', {'message': 'Usuario no registrado en la tabla Usuarios'})

        avisos = Aviso.objects.all()

        context = {
            'estudiante_exists': estudiante_exists,
            'no_control': no_control,
            'avisos': avisos,
        }
        return render(request, "entrevistas/alumno/aviso.html", context)

class informe(View):
    def get(self, request):
        estudiantes = Estudiante.objects.all()
        return render(request, 'entrevistas/administrador/informe.html', {'estudiantes': estudiantes})

class resumenresp(View):
    def get(self, request, idEstudiante):
        estudiante = get_object_or_404(Estudiante, idEstudiante=idEstudiante)
        datos_familiares = get_object_or_404(DatosFamiliares, idEstudiante=estudiante)
        datos_socioeconomicos = get_object_or_404(Socioeconomicos, idEstudiante=estudiante)
        datos_academicos = get_object_or_404(AntecedenteAcademico, idEstudiante=estudiante)
        datos_estudio = get_object_or_404(HabitosEstudio, idEstudiante=estudiante)
        datos_aficiones = get_object_or_404(DatosAficiones,idEstudiante=estudiante)
        datos_personalidad = get_object_or_404(DatosPersonalidad,idEstudiante=estudiante)
        datos_salud = get_object_or_404(DatosSalud,idEstudiante=estudiante)

        context = {
            'estudiante': estudiante,
            'datos_familiares': datos_familiares,
            'datos_socioeconomicos': datos_socioeconomicos,
            'datos_academicos': datos_academicos,
            'datos_estudio': datos_estudio,
            'datos_aficiones':datos_aficiones,
            'datos_personalidad':datos_personalidad,
            'datos_salud':datos_salud
        }

        return render(request, 'entrevistas/alumno/resumenresp.html', context)


def crearEstudiante(request):
    if request.method == 'POST':
        # Obtener los datos del formulario
        carrera = request.POST.get('carrera')
        tutor = request.POST.get('tutor')
        grupo = request.POST.get('grupo')
        fechaR = request.POST.get('fechaR')
        nombre = request.POST.get('nombre')
        noControl = request.POST.get('noControl')
        telCasa = request.POST.get('telCasa')
        correo = request.POST.get('correo')
        telCelular = request.POST.get('telCelular')
        edad = request.POST.get('edad')


        # Crear el estudiante
        try:
            estudiante = Estudiante.objects.create(
                carrera=carrera,
                tutor=tutor,
                grupo=grupo,
                fechaR=fechaR,
                nombre=nombre,
                noControl=noControl,
                telCasa=telCasa,
                correo=correo,
                telCelular=telCelular,
                edad=edad
            )
        except IntegrityError:
            # Mostrar un mensaje de error si ocurre una excepción de integridad de datos
            return render(request, 'entrevistas/alumno/formulario.html', {
                'error_message': 'Los datos ya están registrados. Por favor, verifica la información e intenta nuevamente.'
            })


        # Obtener los datos del familiares
        conviveCon = request.POST.get('conviveCon')
        otra_situacion = request.POST.get('otra_situacion')
        huerfano = request.POST.get('huerfano')
        huerfanoQuien = request.POST.get('huerfanoQuien')
        totalHermanos = request.POST.get('totalHermanos')
        lugarHermano = request.POST.get('lugarHermano')
        estadoCivil = request.POST.get('estadoCivil')

        _idestudiante = estudiante.idEstudiante
        # Crear los datos familiares asociados al estudiante encontrado
        datos_familiares = DatosFamiliares.objects.create(
            idEstudiante= Estudiante.objects.get(idEstudiante = estudiante.idEstudiante),
            conviveCon=conviveCon,
            otra_situacion=otra_situacion,
            huerfano=huerfano,
            huerfanoQuien=huerfanoQuien,
            totalHermanos=totalHermanos,
            lugarHermano=lugarHermano,
            estadoCivil=estadoCivil
        )

         # Obtener los datos socioeconomicos
        vivienda = request.POST.get('vivienda')
        tenencia = request.POST.get('tenencia')
        habitaciones = request.POST.get('habitaciones')
        vehiculos = request.POST.get('vehiculos')
        trabajoPapa = request.POST.get('trabajoPapa')
        trabajoextraPapa = request.POST.get('trabajoextraPapa')
        trabajoMama = request.POST.get('trabajoMama')
        trabajoextraMama = request.POST.get('trabajoextraMama')
        manutencion = request.POST.get('manutencion')
        trabajas = request.POST.get('trabajas')
        relacionCarrera = request.POST.get('relacionCarrera')
        ingresoFam = request.POST.get('ingresoFam')
        contribuyentes = request.POST.get('contribuyentes')
        dependientesIngresos = request.POST.get('dependientesIngresos')


        _idestudiante = estudiante.idEstudiante
        # Crear los datos socioeconomicos asociados al estudiante encontrado
        datos_socioeconomicos = Socioeconomicos.objects.create(
            idEstudiante= Estudiante.objects.get(idEstudiante = estudiante.idEstudiante),
            vivienda=vivienda,
            tenencia=tenencia,
            habitaciones=habitaciones,
            vehiculos=vehiculos,
            trabajoPapa=trabajoPapa,
            trabajoextraPapa=trabajoextraPapa,
            trabajoMama=trabajoMama,
            trabajoextraMama=trabajoextraMama,
            manutencion=manutencion,
            trabajas=trabajas,
            relacionCarrera=relacionCarrera,
            ingresoFam=ingresoFam,
            contribuyentes=contribuyentes,
            dependientesIngresos=dependientesIngresos
        )

         # Obtener los datos academicos
        bachillerato = request.POST.get('bachillerato')
        modalidad = request.POST.get('modalidad')
        duracion = request.POST.get('duracion')
        promedio = request.POST.get('promedio')
        rendimiento = request.POST.get('rendimiento')
        materiasFacil = request.POST.get('materiasFacil')
        materiasDificil = request.POST.get('materiasDificil')
        materiasExtras = request.POST.get('materiasExtras')
        cualesExtras = request.POST.get('cualesExtras')
        repAnio = request.POST.get('repAnio')
        nivelRep = request.POST.get('nivelRep')
        obstaculos = request.POST.get('obstaculos')


        _idestudiante = estudiante.idEstudiante
        # Crear los datos antencedentes academicos asociados al estudiante encontrado
        datos_academicos = AntecedenteAcademico.objects.create(
            idEstudiante= Estudiante.objects.get(idEstudiante = estudiante.idEstudiante),
            bachillerato=bachillerato,
            modalidad=modalidad,
            duracion=duracion,
            promedio=promedio,
            rendimiento=rendimiento,
            materiasFacil=materiasFacil,
            materiasDificil=materiasDificil,
            materiasExtras=materiasExtras,
            cualesExtras=cualesExtras,
            repAnio=repAnio,
            nivelRep=nivelRep,
            obstaculos=obstaculos
        )
         # Obtener los datos estudio
        gustoLectura = request.POST.get('gustoLectura')
        tipoLectura = request.POST.get('tipoLectura')
        sitioLectura = request.POST.get('sitioLectura')
        descripEstudio = request.POST.get('descripEstudio')
        horas = request.POST.get('horas')
        horario = request.POST.get('horario')
        musica = request.POST.get('musica')



        _idestudiante = estudiante.idEstudiante
        # Crear los datos habitos de estudio asociados al estudiante encontrado
        datos_estudio = HabitosEstudio.objects.create(
            idEstudiante= Estudiante.objects.get(idEstudiante = estudiante.idEstudiante),
            gustoLectura=gustoLectura,
            tipoLectura=tipoLectura,
            sitioLectura=sitioLectura,
            descripEstudio=descripEstudio,
            horas=horas,
            horario=horario,
            musica=musica
        )

         # Obtener los datos aficiones
        tiempoLibre = request.POST.get('tiempoLibre')
        horasLibre = request.POST.get('horasLibre')


        _idestudiante = estudiante.idEstudiante
        # Crear los datos aficiones asociados al estudiante encontrado
        datos_aficiones = DatosAficiones.objects.create(
            idEstudiante= Estudiante.objects.get(idEstudiante = estudiante.idEstudiante),
            tiempoLibre=tiempoLibre,
            horasLibre=horasLibre
        )

         # Obtener los datos salud
        estadoSalud = request.POST.get('estadoSalud')
        enfermedadGrave = request.POST.get('enfermedadGrave')
        tipoEnfermedad = request.POST.get('tipoEnfermedad')
        padecimientoCro = request.POST.get('padecimientoCro')
        tipoPadecimiento = request.POST.get('tipoPadecimiento')
        operaciones = request.POST.get('operaciones')
        causaOperacion = request.POST.get('causaOperacion')
        problemaSalud = request.POST.get('problemaSalud')
        condiciones = request.POST.get('condiciones')
        tipoCondiciones = request.POST.get('tipoCondiciones')
        comentario = request.POST.get('comentario')

        _idestudiante = estudiante.idEstudiante
        # Crear los datos salud asociados al estudiante encontrado
        datos_salud = DatosSalud.objects.create(
            idEstudiante= Estudiante.objects.get(idEstudiante = estudiante.idEstudiante),
            estadoSalud=estadoSalud,
            enfermedadGrave=enfermedadGrave,
            tipoEnfermedad=tipoEnfermedad,
            padecimientoCro=padecimientoCro,
            tipoPadecimiento=tipoPadecimiento,
            operaciones=operaciones,
            causaOperacion=causaOperacion,
            problemaSalud=problemaSalud,
            condiciones=condiciones,
            tipoCondiciones=tipoCondiciones,
            comentario=comentario

        )
         # Obtener los datos personalidad
        colaborasCasa = request.POST.get('colaborasCasa')
        colaboracion = request.POST.get('colaboracion')
        ambienteFamiliar = request.POST.get('ambienteFamiliar')
        hablaFamiliar = request.POST.get('hablaFamiliar')
        comunicacion = request.POST.get('comunicacion')

        _idestudiante = estudiante.idEstudiante
        # Crear los datos personalidad asociados al estudiante encontrado
        datos_personalidad = DatosPersonalidad.objects.create(
            idEstudiante= Estudiante.objects.get(idEstudiante = estudiante.idEstudiante),
            colaborasCasa=colaborasCasa,
            colaboracion=colaboracion,
            ambienteFamiliar=ambienteFamiliar,
            hablaFamiliar=hablaFamiliar,
            comunicacion=comunicacion,
        )

        return redirect('resumenresp', idEstudiante=estudiante.idEstudiante)
#Aqui termina parte de entrevistas
class prueba2(View):
   def get(self, request):
        return render(request, 'test/prueba2.html')

def validar_email(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        print(email)
        if User.objects.filter(email=email).exists():
            return JsonResponse({'exists': True})
        else:
            return JsonResponse({'exists': False})
    return JsonResponse({}, status=400)

def login_sso(request):
    if request.method == 'POST':
        email = request.POST.get('email2')
        password = request.POST.get('pswd')

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return render(request, 'login.html', {'Error': 'Correo electrónico no registrado', 'Email': email})

        user_authenticated = authenticate(request, username=user.username, password=password)

        if user_authenticated is not None:
            login(request, user_authenticated)

            # Crear token JWT
            payload = {
                'user_id': user_authenticated.id,
                'exp': datetime.datetime.utcnow() + datetime.timedelta(days=1),
                'iat': datetime.datetime.utcnow()
            }
            token = jwt.encode(payload, settings.SECRET_KEY, algorithm='HS256')

            # Guardar el token en la base de datos
            try:
                usuario = Usuarios.objects.get(User=user_authenticated)
                usuario.token = token
                usuario.save()
            except Usuarios.DoesNotExist:
                return render(request, 'login.html', {'Error': 'Usuario no registrado en la tabla Usuarios', 'Email': email})

            # Establecer una cookie con el token

            tipo=request.user.usuarios.tipo.tipo
            if tipo == "admin":
                response = redirect(reverse('Calendario'))
            elif tipo == "estudiante":
                no_control_usuario = request.user.usuarios.noControl
                existe = Estudiante.objects.filter(noControl=no_control_usuario).exists()
                if existe == True:
                    response = redirect(reverse('aviso'))
                else:
                    response = redirect(reverse('formulario'))
            elif tipo == "tutor":
                response = redirect(reverse('Dashboard'))
            elif tipo == "director" or tipo == "encargadotutorias" or tipo == "directorcarrera":
                response = redirect(reverse('index'))

            elif tipo != "admin" or tipo != "estudiante" or tipo != "tutor" or tipo != "director" or tipo != "encargadotutorias":
                response = redirect(reverse('ResultadosCanalizacion'))

            response.set_cookie('sso_token', token, httponly=False)  # httponly=False para que sea accesible por JavaScript
            return response
        else:
            return render(request, 'login.html', {'Error': 'Credenciales incorrectas', 'Email': email})

    return render(request, 'login.html')

def logout_view(request):
    # Eliminar el token del localStorage en el cliente (JavaScript)
    response = redirect('inicio')  # Ajusta 'login' a la URL de tu página de inicio de sesión
    response.delete_cookie('sso_token')

    # Eliminar el token de la base de datos
    if request.user.is_authenticated:
        try:
            usuario = Usuarios.objects.get(User=request.user)
            usuario.token = None  # O elimina el token de la base de datos según tu modelo
            usuario.save()
        except Usuarios.DoesNotExist:
            pass  # Maneja la excepción según sea necesario

    logout(request)
    return response

def validate_token(request):
    if request.method == 'POST':
        token = request.POST.get('token')

        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            user_id = payload['user_id']
            exp = payload['exp']

            # Verificar si el usuario y el token existen y son válidos
            try:
                usuario = Usuarios.objects.get(User__id=user_id, token=token)
                # Verificar si el token JWT no ha expirado
                if exp > datetime.datetime.utcnow().timestamp():
                    print("Entro")
                    user_id = payload.get('user_id')
                    user = User.objects.get(pk=user_id)
                    tipo=request.user.usuarios.tipo.tipo
                    if tipo == "estudiante":
                        no_control_usuario = request.user.usuarios.noControl
                        existe = Estudiante.objects.filter(noControl=no_control_usuario).exists()
                    else:
                        existe = False
                    return JsonResponse({'valid': True, 'tipo':tipo, 'alumno':existe})
                else:
                    logout(request)
                    return JsonResponse({'valid': False})
            except Usuarios.DoesNotExist:
                logout(request)
                return JsonResponse({'valid': False})

        except jwt.ExpiredSignatureError:
            logout(request)
            return JsonResponse({'valid': False, 'error': 'Token expirado'})
        except jwt.DecodeError:
            logout(request)
            return JsonResponse({'valid': False, 'error': 'Token inválido'})

    return JsonResponse({'error': 'Método no permitido'}, status=405)


class canalizacionIndex(View):
   def get(self, request):
      return render(request,'Canalizacion/index.html')

class canalizacionCalendario(View):

    def get(self, request):
        tipo = request.user.usuarios.tipo.tipo
        grupo = request.user.usuarios.grupo
        canalizaciones = 0
        canalizaciones_mes = 0
        areas = {'psicologo': 'Psicología', 'pedagogo': 'Pedagogía', 'becas': 'Becas', 'enfermeria': 'Enfermería',
                 'incubadora': 'Incubadora', 'bolsadetrabajo': 'Bolsa de trabajo', 'asesoracademico': 'Asesor académico'}

        end_of_today = timezone.now().replace(hour=23, minute=59, second=59, microsecond=999999)
        start_of_today = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)

        if tipo == 'tutor':
            canalizaciones = Canalizacion.objects.filter(
                FechaInicio__isnull=False,
                FechaFinal__isnull=False,
                atencionIndividual__estudiante__grupo=grupo
            )
            canalizaciones_mes = Canalizacion.objects.filter(
                FechaInicio__isnull=False,
                FechaFinal__isnull=False,
                FechaInicio__gte=start_of_today,
                FechaInicio__lte=end_of_today,
                atencionIndividual__estudiante__grupo=grupo
            ).order_by("FechaInicio")
        elif tipo == 'admin':
            canalizaciones = Canalizacion.objects.filter(
                FechaInicio__isnull=False,
                FechaFinal__isnull=False,
            )
            canalizaciones_mes = Canalizacion.objects.filter(
                FechaInicio__isnull=False,
                FechaFinal__isnull=False,
                FechaInicio__gte=start_of_today,
                FechaInicio__lte=end_of_today,
            ).order_by("FechaInicio")
        else:
            canalizaciones = Canalizacion.objects.filter(
                FechaInicio__isnull=False,
                FechaFinal__isnull=False,
                area=areas[tipo]
            )
            canalizaciones_mes = Canalizacion.objects.filter(
                FechaInicio__isnull=False,
                FechaFinal__isnull=False,
                FechaInicio__gte=start_of_today,
                FechaInicio__lte=end_of_today,
                area=areas[tipo]
            ).order_by("FechaInicio")

        lista_canalizaciones = []
        for canalizacion in canalizaciones:
            start = timezone.localtime(canalizacion.FechaInicio).strftime("%Y-%m-%dT%H:%M:%S")
            end = timezone.localtime(canalizacion.FechaFinal).strftime("%Y-%m-%dT%H:%M:%S")

            lista_canalizaciones.append({
                'estudiante': {
                    'id': canalizacion.atencionIndividual.estudiante.User_id,
                    'nombre': canalizacion.atencionIndividual.estudiante.User.first_name,
                    'apellido': canalizacion.atencionIndividual.estudiante.User.last_name,
                    'correo': canalizacion.atencionIndividual.estudiante.User.email
                },
                'title': canalizacion.titulo,
                'description': canalizacion.descripcion,
                'observaciones': canalizacion.observaciones,
                'motivo': canalizacion.motivo,
                'fecha': canalizacion.fecha.strftime("%Y-%m-%dT%H:%M:%S"),
                'start': start,
                'end': end
            })

        context = {'canalizaciones': lista_canalizaciones, 'canalizaciones_mes': canalizaciones_mes}

        return render(request, 'Canalizacion/calendario.html', context)

class canalizacionCompletarSesion(View):
    def get(self, request, id):
        try:
            canalizacion = Canalizacion.objects.filter(atencionIndividual__estudiante=id).order_by('-fecha').first()

            context = {
                'id': id,
                'datos': canalizacion
            }
        except Canalizacion.DoesNotExist:
            context = {
                'id': id,
            }

        return render(request, 'Canalizacion/completarSesion.html', context)

class viewCanalizacionCompletarSesion(View):
    def get(self, request, id):
        datos = get_object_or_404(Canalizacion, id=id)

        context = {
            'id': id,
            'datos': datos
        }

        return render(request, 'Canalizacion/viewCompletarSesion.html', context)

def canalizacionFormCompletarSesion(request, id):
    if request.method == "POST":
        datos = get_object_or_404(Canalizacion, id=id)

        if datos:

            datos.detalles = request.POST['detalles']
            datos.estadoCanalizados = 3
            datos.save()

            estudiante_id = datos.atencionIndividual.estudiante.id
            usuario = Usuarios.objects.get(id=estudiante_id)
            usuario.estado = 1
            usuario.save()

            return redirect('ResultadosCanalizacion')


class viewCanalizar(View):
    def get(self, request, id):
        datos = get_object_or_404(Canalizacion, id=id)

        context = {
            'id': id,
            'datos': datos
        }

        return render(request, 'Canalizacion/viewCanalizar.html', context)

class formCalendario(View):
   def get(self, request, id):
      return render(request,'Canalizacion/formCalendario.html', {'id': id})

def canalizacionFormCalendario(request, id):
    if request.method == "POST":
        datos = Canalizacion.objects.filter(id=id).first()

        datos.titulo = request.POST.get('titulo', datos.titulo)
        datos.descripcion = request.POST.get('descripcion', datos.descripcion)
        datos.FechaInicio = request.POST.get('fechaInicio', datos.FechaInicio)
        datos.FechaFinal = request.POST.get('fechaFinal', datos.FechaFinal)
        datos.estadoCanalizados = 2
        datos.save()

        return redirect('ResultadosCanalizacion')

class canalizacionBajas(View):
    def get(self, request, id):
        try:
            baja = BajaAlumnos.objects.get(estudiante=id)
            estudiante = Usuarios.objects.get(id=id)
            context = {
                'id': id,
                'baja': baja,
                'estudiante': estudiante
            }
        except BajaAlumnos.DoesNotExist:
            context = {
                'id': id,
            }

        return render(request, 'Canalizacion/formBaja.html', context)

def canalizacionBajaAlumno(request, id):
   if request.method == "POST":
      datos = Usuarios.objects.get(id = id)
      tipo = request.POST['tipo']
      observaciones = request.POST['observaciones']
      motivo = request.POST['motivo']
      cicloActual = Periodo.objects.filter(estado = True).first()

      Baja  = BajaAlumnos.objects.create(tipo=tipo, observaciones=observaciones, motivo=motivo, cicloAccion=cicloActual, estudiante=datos)

      datos.estado = 3
      datos.save()

      return redirect('Dashboard')

class canalizacionFormCanalizar(View):
    def get(self, request, id):
        try:
            atencion_individual = AtencionIndividual.objects.filter(estudiante=id).order_by('-fecha').first()

            canalizacion = Canalizacion.objects.filter(atencionIndividual=atencion_individual).first()

            context = {
                'id': id,
                'canalizacion': canalizacion,
                'atencion_individual': atencion_individual
            }
        except (AtencionIndividual.DoesNotExist, Canalizacion.DoesNotExist):
            context = {
                'id': id,
            }

        return render(request, 'Canalizacion/formCanalizar.html', context)

def canalizacionFormCanalizarAlumno(request, id):
    if request.method == "POST":

        atencion_individual = AtencionIndividual.objects.filter(estudiante=id).order_by('-fecha').first()

        area = request.POST.get('area')
        observaciones = request.POST.get('observaciones')
        motivo = request.POST.get('motivo')
        ciclo_actual = Periodo.objects.filter(estado=True).first()
        Canalizar = Canalizacion.objects.create(
            area=area,
            observaciones=observaciones,
            motivo=motivo,
            cicloAccion=ciclo_actual,
            atencionIndividual=atencion_individual,
            estadoCanalizados=1
        )

        usuario = atencion_individual.estudiante
        usuario.estado = 4
        usuario.save()

        return redirect('Dashboard')

class canalizacionFormCerrarTutorias(View):
   def get(self, request):
      return render(request,'Canalizacion/formCerrarTutorias.html')

def cerrarTurorias(request):
   if request.method == "POST":
      cierreTutorias = request.POST['cierreTutorias']
      cicloActual = Periodo.objects.filter(estado = True)
      userActual = request.user
      cierre  = CierreTutorias.objects.create(cierreTutorias=cierreTutorias, cicloAccion=cicloActual[0], tutor=userActual)
      cierre.save()
      return redirect('Dashboard')

class canalizacionReportes(View):
    def get(self, request, id):
        if request.GET.get('periodo_id'):
            return self.get_ajax_response(request, id)
        else:
            return self.get_page_response(request, id)

    def get_page_response(self, request, id):
        periodos = Periodo.objects.all().order_by('-id')
        alumno = Usuarios.objects.select_related('User').get(id=id)
        canalizaciones = Canalizacion.objects.filter(atencionIndividual__estudiante=id).select_related('atencionIndividual', 'atencionIndividual__estudiante')

        reportes_data = []
        for canalizacion in canalizaciones:
            reportes_data.append({
                'area': canalizacion.area,
                'nombre': canalizacion.atencionIndividual.estudiante.User.first_name,
                'apellidos': canalizacion.atencionIndividual.estudiante.User.last_name,
                'no_control': canalizacion.atencionIndividual.estudiante.noControl,
                'asunto_atencion': canalizacion.motivo,
                'observaciones': canalizacion.observaciones,
                'detalles': canalizacion.detalles,
                'fecha': canalizacion.atencionIndividual.fecha,
            })

        context = {
            'reportes_data': reportes_data,
            'periodos': periodos,
            'alumno': alumno,
            'alumno_id': id
        }

        return render(request, 'Canalizacion/Reportes.html', context)

    def get_ajax_response(self, request, id):
        periodo_id = request.GET.get('periodo_id')

        if periodo_id:
            canalizaciones = Canalizacion.objects.filter(cicloAccion_id=periodo_id, atencionIndividual__estudiante=id).values(
                'area',
                'atencionIndividual__estudiante__User__first_name',
                'atencionIndividual__estudiante__User__last_name',
                'atencionIndividual__estudiante__noControl',
                'motivo',
                'observaciones',
                'detalles',
                'atencionIndividual__fecha'
            )

            reportes_data = list(canalizaciones)
            data = {'reportes_data': reportes_data}
            return JsonResponse(data)

        return JsonResponse({'error': 'Periodo no válido'}, status=400)

class generatePDF(View):
    def get(self, request, id):
        alumno = get_object_or_404(Usuarios, id=id)
        canalizaciones = Canalizacion.objects.filter(atencionIndividual__estudiante_id=id).select_related('atencionIndividual', 'atencionIndividual__estudiante')

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_{alumno.User.first_name}_{alumno.User.last_name}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=letter)
        elements = []

        logo_path = finders.find('../statics/assets/img/utem.png')
        elements.append(ImageCanalizacion(logo_path, 1*inch, 1*inch))

        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f'Reporte de Canalizaciones para {alumno.User.first_name} {alumno.User.last_name}',
                                  getSampleStyleSheet()['Title']))

        styles = getSampleStyleSheet()
        normal_style = styles['Normal']
        normal_style.wordWrap = 'CJK'

        header_style = styles['Normal'].clone('header')
        header_style.textColor = colors.whitesmoke
        header_style.fontName = 'Helvetica-Bold'

        data = [
            [
                Paragraph('Area', header_style),
                Paragraph('Nombre', header_style),
                Paragraph('Apellidos', header_style),
                Paragraph('No. de Control', header_style),
                Paragraph('Asunto', header_style),
                Paragraph('Observaciones', header_style),
                Paragraph('Detalles', header_style),
                Paragraph('Fecha', header_style)
            ]
        ]

        for canalizacion in canalizaciones:
            data.append([
                Paragraph(canalizacion.area, normal_style),
                Paragraph(canalizacion.atencionIndividual.estudiante.User.first_name, normal_style),
                Paragraph(canalizacion.atencionIndividual.estudiante.User.last_name, normal_style),
                Paragraph(str(canalizacion.atencionIndividual.estudiante.noControl), normal_style),
                Paragraph(canalizacion.motivo, normal_style),
                Paragraph(canalizacion.observaciones, normal_style),
                Paragraph(canalizacion.detalles, normal_style),
                Paragraph(canalizacion.atencionIndividual.fecha.strftime('%d-%m-%Y'), normal_style)
            ])

        col_widths = [50, 70, 70, 60, 80, 100, 100, 60]

        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        elements.append(table)

        doc.build(elements)

        return response

class canalizacionIndex(View):
   def get(self, request):
      tutor = request.user
      if tutor.usuarios.grupo == None:
        estudiantes = Usuarios.objects.filter(tipo__tipo='estudiante')
        periodo = Periodo.objects.filter(estado=1)
        context = {
            'TablaViews': estudiantes,
            'Periodo': periodo

        }
      else:
        estudiantes = Usuarios.objects.filter(tipo__tipo='estudiante', grupo=tutor.usuarios.grupo)
        periodo = Periodo.objects.filter(estado=1)
        context = {
            'TablaViews': estudiantes,
            'Periodo': periodo
        }

      return render(request, 'Canalizacion/index.html', context)


class canalizacionExpedientes(View):
    def get(self, request, id):
        if request.GET.get('periodo_id'):
            return self.get_ajax_response(request, id)
        else:
            return self.get_page_response(request, id)

    def get_page_response(self, request, id):
        periodos = Periodo.objects.all().order_by('-id')
        alumno = id

        atencion_ids = AtencionIndividual.objects.filter(estudiante_id=alumno).values_list('id', flat=True)

        TablaExpedientes = Canalizacion.objects.filter(atencionIndividual_id__in=atencion_ids).annotate(
            observacionesIndividual=F('atencionIndividual__observaciones'),
            asuntoTratarIndividual=F('atencionIndividual__asuntoTratar'),
            fechaIndividual=F('atencionIndividual__fecha'),
        )

        alumno_exp = Usuarios.objects.filter(id=alumno)

        # Formatear las fechas antes de pasarlas al contexto usando strftime
        for exp in TablaExpedientes:
            exp.fechaIndividual = exp.fechaIndividual.strftime('%d-%m-%Y a las %H:%M')
            exp.fecha = exp.fecha.strftime('%d-%m-%Y a las %H:%M')

        context = {
            'periodos': periodos,
            'TablaExpedientes': TablaExpedientes,
            'alumno': alumno_exp,
            'alumno_id': alumno,
            'id': id
        }

        return render(request, 'Canalizacion/expediente.html', context)

    def get_ajax_response(self, request, id):
        alumno = request.GET.get('alumno')
        periodo_id = request.GET.get('periodo_id')

        atencion_ids = AtencionIndividual.objects.filter(estudiante_id=alumno).values_list('id', flat=True)

        if periodo_id:
            atenciones_individuales = Canalizacion.objects.filter(
                cicloAccion_id=periodo_id,
                atencionIndividual_id__in=atencion_ids
            ).values(
                'area',
                'atencionIndividual__observaciones',
                'atencionIndividual__asuntoTratar',
                'atencionIndividual__fecha'
            )

            canalizaciones = Canalizacion.objects.filter(
                cicloAccion_id=periodo_id,
                atencionIndividual_id__in=atencion_ids
            ).values(
                'area',
                'detalles',
                'fecha'
            )

            # Formatear las fechas antes de enviarlas usando strftime
            for atencion in atenciones_individuales:
                atencion['atencionIndividual__fecha'] = atencion['atencionIndividual__fecha'].strftime('%d-%m-%Y a las %H:%M')
            for canalizacion in canalizaciones:
                canalizacion['fecha'] = canalizacion['fecha'].strftime('%d-%m-%Y a las %H:%M')

            data = {
                'atenciones_individuales': list(atenciones_individuales),
                'canalizaciones': list(canalizaciones),
                'id': id
            }

            return JsonResponse(data)

        return JsonResponse({'error': 'Periodo no válido'}, status=400)


class canalizacionResultadosCanalizacion(View):
   def get(self, request):
      user = request.user.usuarios.tipo.tipo

      areas = {'psicologo':'Psicología', 'pedagogo':'Pedagogía', 'becas':'Becas', 'enfermeria':'Enfermería', 'incubadora':'Incubadora', 'bolsadetrabajo':'Bolsa de trabajo', 'asesoracademico':'Asesor académico'}

      if user=='admin':
        tabla = Canalizacion.objects.select_related('atencionIndividual').order_by('-fecha')
        context = {
            'TablaResultados': tabla,
        }
      else:
        tabla = Canalizacion.objects.select_related('atencionIndividual').filter(area=areas[user]).order_by('-fecha')
        context = {
            'TablaResultados': tabla,
        }
      return render(request,'Canalizacion/resultadosCanalizacion.html', context)


def actividadTutorial(request):
    actividades = AccionTutorial.objects.filter(cicloAccion__estado=True, tutor=request.user)
    messages.success(request, '¡Datos cargados!')
    return render(request, 'documentos/actividadTutorial.html', {"actividades":actividades})


def eliminarActividadTutorial(request,id):
    actividad = get_object_or_404(AccionTutorial, id=id)
    actividad.delete()
    messages.success(request, '¡Actividad eliminada!')
    return redirect('actividadTutorial')

def agregarActividadTutorial(request):
    if request.method == 'POST':
        tema = request.POST.get('tema')
        objetivos = request.POST.get('objetivos')
        actividades = request.POST.get('actividades')
        recursos = request.POST.get('recursos')
        evidencias = request.POST.get('evidencias')

        # Obtén el único Periodo con estado=True
        try:
            cicloAccion = Periodo.objects.get(estado=True)
        except Periodo.DoesNotExist:
            return HttpResponse("Error: No hay un periodo activo.")

        # Obtén el usuario logueado como tutor
        tutor = request.user

        # Crea una nueva instancia de AccionTutorial y guárdala
        nueva_accion = AccionTutorial(
            tema=tema,
            objetivos=objetivos,
            actividades=actividades,
            recursos=recursos,
            cicloAccion=cicloAccion,
            evidencias = evidencias,
            tutor = tutor
        )
        nueva_accion.save()
        messages.success(request, '¡Guardado con éxito!')
        return redirect('actividadTutorial')  # Redirige a la página de actividades después de guardar

    return render(request, 'documentos/agregarActividad.html')

def editarActividadTutorial(request, id):
    actividad = get_object_or_404(AccionTutorial, id=id)

    if request.method == 'POST':
        # Procesar el formulario enviado
        actividad.tema = request.POST.get('tema')
        actividad.objetivos = request.POST.get('objetivos')
        actividad.recursos = request.POST.get('recursos')
        actividad.actividades = request.POST.get('actividades')
        actividad.evidencias = request.POST.get('evidencias')
        actividad.save()

        messages.success(request, '¡Guardado con éxito!')
        return redirect('editarActividad', id=actividad.id)

    # Si es un GET, simplemente renderiza el formulario
    return render(request, 'documentos/editarActividad.html', {'actividad': actividad})

def infoActividadTutorial(request, id):
    actividad = get_object_or_404(AccionTutorial, cicloAccion__estado=True, tutor=request.user, id=id)
    return render(request,'documentos/infoActividad.html', {"actividad":actividad})

def infoatencionIndividual(request):
    tutor = request.user
    usuario = Usuarios.objects.get(User=tutor)
    grupo= usuario.grupo
    atenciones = AtencionIndividual.objects.filter(cicloAccion__estado=True, estudiante__grupo=grupo)
    messages.success(request, '¡Datos cargados!')
    return render(request, 'documentos/atencionIndividual.html', {"atenciones": atenciones})

def agregarAtencionIndividual(request):
    periodo_activo = Periodo.objects.filter(estado=True).first()
    if request.method == 'POST':
        estudiante_id = request.POST.get('estudianteAtencion')
        asuntoTratar = request.POST.get('asuntoTratar')
        observaciones = request.POST.get('observaciones')


        # Crear la instancia de AtencionIndividual
        AtencionIndividual.objects.create(
            estudiante_id=estudiante_id,
            asuntoTratar=asuntoTratar,
            observaciones=observaciones,
            cicloAccion=periodo_activo

        )

        cambioEstado=Usuarios.objects.get(id=estudiante_id)
        cambioEstado.estado= 2
        cambioEstado.save()


        # cambioEstado.save()

        messages.success(request, '¡Atención individual registrada con éxito!')
        return redirect('atencionIndividual')

    # Filtrar estudiantes que pertenecen al grupo del tutor actual y que son de tipo 'estudiante'
    tutor = request.user
    estudiantes = Usuarios.objects.filter(tipo__tipo='estudiante', grupo=tutor.usuarios.grupo)

    return render(request, 'documentos/registrarAtencion.html', {'estudiantes': estudiantes})


def editarAtencionIndividual(request, id):
    atencion = get_object_or_404(AtencionIndividual, id=id)
    tutor = request.user
    estudiantes = Usuarios.objects.filter(tipo__tipo='estudiante', grupo=tutor.usuarios.grupo)


    if request.method == 'POST':
        atencion.asuntoTratar = request.POST.get('asuntoTratar')
        atencion.observaciones = request.POST.get('observaciones')
        atencion.save()

        messages.success(request, '¡Atención individual editada con éxito!')
        return redirect('editarAtencion', id=atencion.id)

    return render(request, 'documentos/editarAtencion.html', {'atencion': atencion, 'estudiantes': estudiantes})

def eliminarAtencionIndividual(request, id):
    atencion = get_object_or_404(AtencionIndividual, id=id)
    usuarioId=atencion.estudiante.id
    cambioEstado=Usuarios.objects.get(id=usuarioId)
    cambioEstado.estado= 1
    cambioEstado.save()

    atencion.delete()
    messages.success(request, '¡Atención individual eliminada con éxito!')
    return redirect('atencionIndividual')




def infoeditarAtencion(request):
    return render (request,'documentos/editarAtencion.html')

def inforegistrarAtencion(request):
    return render (request, 'documentos/registrarAtencion.html')

def informePlanAccion(request):
    return render (request, 'documentos/informePlanAccion.html')

def reportePlanAccion(request):
    return render (request, 'documentos/reportePlanAccion.html')

def evaluacionAcTutorial(request):
    nombre_maestro = None  # Inicializar la variable nombre_maestro
    estudiante = request.user
    grupo_estudiante = estudiante.usuarios.grupo


    maestro = Usuarios.objects.filter(
                tipo__tipo='tutor',  # Ajusta según tu modelo TipoUsuario
                grupo=grupo_estudiante
            ).first()

    if maestro:
                nombre_maestro = maestro.User.get_full_name()
    else:
                nombre_maestro = "Tutor no encontrado"  # Manejo de caso sin maestro

    periodo_activo = Periodo.objects.filter(estado=True).first()

    if not periodo_activo:
            messages.error(request, 'No hay un período activo disponible.')
            return render(request, 'documentos/evaluacionAcTutorial.html')

        # Verificar si el usuario ya ha contestado la encuesta para el período activo
    evaluacion_existente = EvaluacionTutor.objects.filter(
            estudiante=estudiante,
            cicloEvaluacion=periodo_activo
        ).exists()

    if evaluacion_existente:
            messages.error(request, 'Ya has contestado la encuesta previamente.')
            return render(request, 'documentos/evaluacionAcTutorial.html', {'nombre_maestro': nombre_maestro})

    if request.method == 'POST':
        # Guardar la evaluación del tutor
        evaluacion = EvaluacionTutor(
            estudiante=estudiante,
            puntualidad=request.POST.get('puntualidad'),
            proposito=request.POST.get('proposito'),
            planTrabajo=request.POST.get('planTrabajo'),
            temasPrevistos=request.POST.get('temasPrevistos'),
            temasInteres=request.POST.get('temasInteres'),
            disposicionTutor=request.POST.get('disposicionTutor'),
            cordialidad=request.POST.get('cordialidad'),
            orientacion=request.POST.get('orientacion'),
            dominio=request.POST.get('dominio'),
            impacto=request.POST.get('impacto'),
            serviciosApoyo=request.POST.get('serviciosApoyo'),
            cicloEvaluacion=periodo_activo
        )
        if evaluacion_existente:
            messages.error(request, 'Ya has contestado la encuesta previamente.')
            return render(request, 'documentos/evaluacionAcTutorial.html', {'nombre_maestro': nombre_maestro})
        else:
            evaluacion.save()


        messages.success(request, '¡Encuesta guardada con éxito!')
        return render(request, 'documentos/evaluacionAcTutorial.html', {'nombre_maestro': nombre_maestro})

    # Si es un método GET o cualquier otro, simplemente renderizar el formulario
    return render(request, 'documentos/evaluacionAcTutorial.html', {'nombre_maestro': nombre_maestro})



def descargarXLSX(request):
    tutor = request.user
    usuario = Usuarios.objects.get(User=tutor)
    periodo_activo = Periodo.objects.filter(estado=True).first()
    grupo= usuario.grupo
    atenciones= AtencionIndividual.objects.filter(cicloAccion__estado=True, estudiante__grupo=grupo)
    actividad = AccionTutorial.objects.filter(tutor=tutor, cicloAccion=periodo_activo)
# Cargar el archivo base
    path_archivo_base = os.path.join(settings.BASE_DIR, 'mainApp', 'data', 'basePlanAccion.xlsx')
    # Cargar el archivo base
    wb = load_workbook(filename=path_archivo_base, read_only=False)
    ws = wb.active

    if periodo_activo:
        ws['F6'] = f"{periodo_activo.periodo} {periodo_activo.anio}"
    else:
        ws['F6'] = "No hay periodo activo"

    ws['C5'] = f"{tutor.first_name} {tutor.last_name}"
    ws['C6'] = usuario.grupo

    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    # Estilo de fuente
    calibri_10_font = Font(name='Calibri', size=10)
    calibri_10_font_bold = Font(name='Calibri', size=10, bold=True)
    gray_fill = PatternFill(start_color='D8D8D8', end_color='D8D8D8', fill_type='solid')

    # Rellenar las filas con los datos de 'actividad'
    row_num = 10
    inserted_rows_count = 0
    for idx, act in enumerate(actividad, start=1):
        inserted_rows_count += 1
        num_cell = ws.cell(row=row_num, column=1)
        num_cell.value = idx
        num_cell.font = calibri_10_font_bold
        num_cell.border = thin_border

        # Llenado de datos en las columnas
        ws.cell(row=row_num, column=2).value = act.tema
        ws.cell(row=row_num, column=3).value = act.objetivos
        ws.cell(row=row_num, column=4).value = act.actividades
        ws.cell(row=row_num, column=5).value = act.recursos
        ws.cell(row=row_num, column=6).value = act.evidencias

        # Aplicar bordes y fuente a cada celda en la fila actual
        for col in range(2, 7):
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
            cell.font = calibri_10_font

        row_num += 1


        # Calcular alumnos_atencion_row y atencion_start_row
    alumnos_atencion_row = max(row_num, 24) + inserted_rows_count -10
    atencion_start_row = alumnos_atencion_row + 1

        # Insertar y combinar la celda "Alumnos a atención individual" dos filas antes de alumnos_atencion_row
    merge_start_row = alumnos_atencion_row - 1
    ws.merge_cells(start_row=merge_start_row, start_column=1, end_row=merge_start_row, end_column=3)
    merged_cell = ws.cell(row=merge_start_row, column=1)
    merged_cell.value = "Alumnos a atención individual"
    merged_cell.font = calibri_10_font_bold
    merged_cell.fill = gray_fill  # Aplicar fondo gris
    merged_cell.border = thin_border  # Aplicar bordes a la celda combinada

# Aplicar el borde a cada celda dentro de la región combinada
    for col in range(1, 4):
        cell = ws.cell(row=merge_start_row, column=col)
        cell.border = thin_border

    num_cell = ws.cell(row=alumnos_atencion_row, column=1)
    num_cell.value = "No."
    num_cell.font = calibri_10_font_bold
    num_cell.border = thin_border
    num_cell.fill = gray_fill

    ws.merge_cells(start_row=alumnos_atencion_row, start_column=2, end_row=alumnos_atencion_row, end_column=3)
    cell = ws.cell(row=alumnos_atencion_row, column=2)
    cell.value =  "Nombre"
    cell.font = calibri_10_font_bold
    cell.border = thin_border
    cell.fill = gray_fill

    asunto_cell = ws.cell(row=alumnos_atencion_row, column=4)
    asunto_cell.value = "Asunto a tratar"
    asunto_cell.font = calibri_10_font_bold
    asunto_cell.border = thin_border
    asunto_cell.fill = gray_fill

    ws.merge_cells(start_row=alumnos_atencion_row, start_column=5, end_row=alumnos_atencion_row, end_column=6)

    # Obtener la celda combinada
    merged_cell = ws.cell(row=alumnos_atencion_row, column=5)

    # Configurar el contenido y estilo de la celda combinada
    merged_cell.value = "Observaciones"
    merged_cell.font = calibri_10_font_bold
    merged_cell.fill = gray_fill  # Aplicar fondo gris
    merged_cell.border = thin_border  # Aplicar bordes a la celda combinada

    # Aplicar el borde a cada celda dentro de la región combinada
    for col in range(5, 7):
        cell = ws.cell(row=alumnos_atencion_row, column=col)
        cell.border = thin_border

        # Rellenar las filas con los datos de 'atenciones'
    for idx, atencion in enumerate(atenciones, start=1):
            # Numeración continua en la columna 1
            num_cell = ws.cell(row=atencion_start_row, column=1)
            num_cell.value = idx
            num_cell.font = calibri_10_font_bold
            num_cell.border = thin_border

            # Combinar columnas 2 y 3
            ws.merge_cells(start_row=atencion_start_row, start_column=2, end_row=atencion_start_row, end_column=3)
            cell = ws.cell(row=atencion_start_row, column=2)
            cell.value =  f"{atencion.estudiante.User.first_name} {atencion.estudiante.User.last_name}"
            cell.font = calibri_10_font
            cell.border = thin_border

            # Aplicar el borde a la celda combinada
            for col in range(2, 4):
                merged_cell = ws.cell(row=atencion_start_row, column=col)
                merged_cell.border = thin_border

            # Llenar datos en las columnas restantes
            asunto_cell = ws.cell(row=atencion_start_row, column=4)
            asunto_cell.value = atencion.asuntoTratar
            asunto_cell.font = calibri_10_font
            asunto_cell.border = thin_border

            ws.merge_cells(start_row=atencion_start_row, start_column=5, end_row=atencion_start_row, end_column=6)
            observaciones_cell = ws.cell(row=atencion_start_row, column=5)
            observaciones_cell.value = atencion.observaciones
            observaciones_cell.font = calibri_10_font
            observaciones_cell.border = thin_border

            # Aplicar el borde a la celda combinada (observaciones)
            for col in range(5, 7):
                merged_cell = ws.cell(row=atencion_start_row, column=col)
                merged_cell.border = thin_border

            atencion_start_row += 1

    ws.insert_rows(atencion_start_row + 2)

    # Encabezados en la fila insertada
    headers = ["No.", "Temas no vistos", "Motivo", "Observaciones"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=atencion_start_row + 2, column=col_idx)
        cell.value = header
        cell.font = calibri_10_font_bold
        cell.border = thin_border
        cell.fill = gray_fill

    # Insertar tres filas adicionales
    for i in range(1, 4):
        ws.insert_rows(atencion_start_row + 3 + i)

        # Llenar celdas en las nuevas filas
        num_cell = ws.cell(row=atencion_start_row + 2 + i, column=1)
        num_cell.value = i
        num_cell.font = calibri_10_font_bold
        num_cell.border = thin_border

        tema_cell = ws.cell(row=atencion_start_row + 2 + i, column=2)
        tema_cell.border = thin_border

        motivo_cell = ws.cell(row=atencion_start_row + 2 + i, column=3)
        motivo_cell.border = thin_border

        ob_cell = ws.cell(row=atencion_start_row + 2 + i, column=4)
        ob_cell.border = thin_border

    # Crear un objeto BytesIO para guardar el archivo modificado en memoria
    with BytesIO() as in_memory_fp:
        wb.save(in_memory_fp)
        in_memory_fp.seek(0)

        # Crear la respuesta HTTP
        response = HttpResponse(in_memory_fp.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=8.- FOR-06-01-B_r1 Plan de Accion Tutorial.xlsx'

    return response


def descargarReporte(request):
    path_archivo_base_dos = os.path.join(settings.BASE_DIR, 'mainApp', 'data', 'baseReportePlanAccion.xlsx')
    wb = load_workbook(filename=path_archivo_base_dos, read_only=False)
    ws = wb.active

    tutor = request.user
    periodo_activo = Periodo.objects.filter(estado=True).first()
    usuario_logueado = request.user
    usuario_obj = Usuarios.objects.get(User=usuario_logueado)
    grupo_usuario_logueado = usuario_obj.grupo


    noProgramadas= AtencionIndividual. objects.filter(cicloAccion=periodo_activo, estudiante__grupo=grupo_usuario_logueado).count()


    dificultadCierre = CierreTutorias.objects.filter(tutor=tutor, cicloAccion=periodo_activo).values_list('cierreTutorias', flat=True)




    area_P = "Pedagogía"
    motivos = Canalizacion.objects.filter(
        area=area_P,
        cicloAccion=periodo_activo,
        atencionIndividual__estudiante__grupo=grupo_usuario_logueado
    ).values_list('motivo', flat=True)

    count_pedagogia = Canalizacion.objects.filter(area=area_P,cicloAccion=periodo_activo, atencionIndividual__estudiante__grupo=grupo_usuario_logueado).count()

    area_Psi = "Psicología"
    motivo_psi = Canalizacion.objects.filter(
        area=area_Psi,
        cicloAccion=periodo_activo,
        atencionIndividual__estudiante__grupo=grupo_usuario_logueado
    ).values_list('motivo', flat=True)
    count_Psicología = Canalizacion.objects.filter(area=area_Psi,cicloAccion=periodo_activo, atencionIndividual__estudiante__grupo=grupo_usuario_logueado).count()

    area_B = "Becas"
    motivo_B = Canalizacion.objects.filter(
        area=area_B,
        cicloAccion=periodo_activo,
        atencionIndividual__estudiante__grupo=grupo_usuario_logueado
    ).values_list('motivo', flat=True)
    count_Becas = Canalizacion.objects.filter(area=area_B,cicloAccion=periodo_activo, atencionIndividual__estudiante__grupo=grupo_usuario_logueado).count()

    area_E = "Enfermería"
    motivo_E = Canalizacion.objects.filter(
        area=area_E,
        cicloAccion=periodo_activo,
        atencionIndividual__estudiante__grupo=grupo_usuario_logueado
    ).values_list('motivo', flat=True)
    count_Enfermería = Canalizacion.objects.filter(area=area_E,cicloAccion=periodo_activo, atencionIndividual__estudiante__grupo=grupo_usuario_logueado).count()

    area_I = "Incubadora"
    motivo_I = Canalizacion.objects.filter(
        area=area_I,
        cicloAccion=periodo_activo,
        atencionIndividual__estudiante__grupo=grupo_usuario_logueado
    ).values_list('motivo', flat=True)
    count_Incubadora =  Canalizacion.objects.filter(area=area_I,cicloAccion=periodo_activo, atencionIndividual__estudiante__grupo=grupo_usuario_logueado).count()


    area_BT = "Bolsa de Trabajo"
    motivo_BT = Canalizacion.objects.filter(
        area=area_BT,
        cicloAccion=periodo_activo,
        atencionIndividual__estudiante__grupo=grupo_usuario_logueado
    ).values_list('motivo', flat=True)
    count_BT =  Canalizacion.objects.filter(area=area_BT,cicloAccion=periodo_activo, atencionIndividual__estudiante__grupo=grupo_usuario_logueado).count()

    area_AA = "Asesor Académico"
    motivo_AA = Canalizacion.objects.filter(
        area=area_AA,
        cicloAccion=periodo_activo,
        atencionIndividual__estudiante__grupo=grupo_usuario_logueado
    ).values_list('motivo', flat=True)
    count_AA = Canalizacion.objects.filter(area=area_AA,cicloAccion=periodo_activo, atencionIndividual__estudiante__grupo=grupo_usuario_logueado).count()


    AsuntoAsesoria = 'Asesoría'
    AsuntoDiciplina = 'Disciplina'
    AsuntoOrientacion = 'Orientación'
    AsuntoAdministrativa = 'Administrativa'
    AsuntoOtra = 'Otra'

    count_Asesoria = AtencionIndividual.objects.filter(estudiante__grupo=grupo_usuario_logueado, cicloAccion=periodo_activo, asuntoTratar=AsuntoAsesoria).count()
    count_Diciplina = AtencionIndividual.objects.filter(estudiante__grupo=grupo_usuario_logueado, cicloAccion=periodo_activo, asuntoTratar=AsuntoDiciplina).count()
    count_Orientacion = AtencionIndividual.objects.filter(estudiante__grupo=grupo_usuario_logueado, cicloAccion=periodo_activo, asuntoTratar=AsuntoOrientacion).count()
    count_Administrativa = AtencionIndividual.objects.filter(estudiante__grupo=grupo_usuario_logueado, cicloAccion=periodo_activo, asuntoTratar=AsuntoAdministrativa).count()
    count_Otra = AtencionIndividual.objects.filter(estudiante__grupo=grupo_usuario_logueado, cicloAccion=periodo_activo, asuntoTratar=AsuntoOtra).count()



    observaciones_Asesoria = AtencionIndividual.objects.filter(estudiante__grupo=grupo_usuario_logueado, cicloAccion=periodo_activo, asuntoTratar=AsuntoAsesoria).values_list('observaciones', flat=True)
    observaciones_Diciplina = AtencionIndividual.objects.filter(estudiante__grupo=grupo_usuario_logueado, cicloAccion=periodo_activo, asuntoTratar=AsuntoDiciplina).values_list('observaciones', flat=True)
    observaciones_Orientacion = AtencionIndividual.objects.filter(estudiante__grupo=grupo_usuario_logueado, cicloAccion=periodo_activo, asuntoTratar=AsuntoOrientacion).values_list('observaciones', flat=True)
    observaciones_Administrativa = AtencionIndividual.objects.filter(estudiante__grupo=grupo_usuario_logueado, cicloAccion=periodo_activo, asuntoTratar=AsuntoAdministrativa).values_list('observaciones', flat=True)
    observaciones_Otra = AtencionIndividual.objects.filter(estudiante__grupo=grupo_usuario_logueado, cicloAccion=periodo_activo, asuntoTratar=AsuntoOtra).values_list('observaciones', flat=True)


    Estado_Realizadas = 1
    Estado_Canalizadas = 2
    Estado_Programadas = 0

    count_EstadoP = Canalizacion.objects.filter(estadoCanalizados = Estado_Programadas).count()
    count_EstadoR = Canalizacion.objects.filter(estadoCanalizados = Estado_Realizadas).count()
    count_EstadoC = Canalizacion.objects.filter(estadoCanalizados = Estado_Canalizadas).count()


    def convert_to_string(motivos_list):
        return ', '.join(motivos_list) if motivos_list else ''

    # Set each variable in a specific cell
    ws['F9'] = convert_to_string(motivo_psi)
    ws['F10'] = convert_to_string(motivos)
    ws['F11'] = convert_to_string(motivo_B)
    ws['F12'] = convert_to_string(motivo_E)
    ws['F13'] = convert_to_string(motivo_I)
    ws['F14'] = convert_to_string(motivo_BT)
    ws['F15'] = convert_to_string(motivo_AA)

    ws['E9'] = count_Psicología
    ws['E10'] = count_pedagogia
    ws['E11'] = count_Becas
    ws['E12'] = count_Enfermería
    ws['E13'] = count_Incubadora
    ws['E14'] = count_BT
    ws['E15'] = count_AA

    ws['A9'] = count_EstadoP
    ws['B9'] = count_EstadoR
    ws['C9'] = count_EstadoC

    ws['A22'] = noProgramadas

    ws['E22'] = count_Asesoria
    ws['E23'] = count_Diciplina
    ws['E24'] = count_Orientacion
    ws['E25'] = count_Administrativa
    ws['E26'] = count_Otra

    if dificultadCierre:
     ws['G23'] = dificultadCierre[0]
    else:
     ws['G23'] = 'No hay comentario'


    ws['F22'] = convert_to_string(observaciones_Asesoria)
    ws['F23'] = convert_to_string(observaciones_Diciplina)
    ws['F24'] = convert_to_string(observaciones_Orientacion)
    ws['F25'] = convert_to_string(observaciones_Administrativa)
    ws['F26'] = convert_to_string(observaciones_Otra)

    ws['D22'] = AsuntoAsesoria
    ws['D23'] = AsuntoDiciplina
    ws['D24'] = AsuntoOrientacion
    ws['D25'] = AsuntoAdministrativa
    ws['D26'] = AsuntoOtra

    # Obtiene todas las bajas de alumnos del grupo del tutor
    bajas = BajaAlumnos.objects.filter(estudiante__usuarios__grupo=grupo_usuario_logueado, cicloAccion=periodo_activo)

    # Diccionario para mapear motivos de baja a filas
    motivos_fila = {
    "No se cumplieron expectativas": 9,
    "Reprobación": 10,
    "Problemas económicos": 11,
    "Dificultades para el transporte": 12,
    "Problemas de trabajo": 13,
    "Cambio de carrera": 14,
    "Incompatibilidad de horario": 15,
    "Faltas al reglamento": 16,
    "Cambio de residencia": 17,
    "Cambio de universidad": 18,
    "Problemas familiares": 19,
    "Problemas personales": 20,
    "Otra": 21,
    }

    # Inicializa los contadores de cantidad y listas para tipos y observaciones
    cantidad_dict = {motivo: 0 for motivo in motivos_fila.keys()}
    tipos_dict = {motivo: [] for motivo in motivos_fila.keys()}
    observaciones_dict = {motivo: [] for motivo in motivos_fila.keys()}

    # Itera sobre las bajas y cuenta las cantidades por motivo
    for baja in bajas:
        fila = motivos_fila.get(baja.motivo)
        if fila is not None:
            cantidad_dict[baja.motivo] += 1
            tipos_dict[baja.motivo].append(baja.tipo)
            observaciones_dict[baja.motivo].append(baja.observaciones)

        ws[f'I{fila}'] = cantidad_dict[baja.motivo]
        ws[f'J{fila}'] = ', '.join(tipos_dict[baja.motivo])
        ws[f'K{fila}'] = ', '.join(observaciones_dict[baja.motivo])

        # Obtén la celda superior izquierda del rango fusionado para observaciones
        merged_cells = ws.merged_cells.ranges
        assigned = False
        for merged_cell in merged_cells:
            if merged_cell.min_row <= fila <= merged_cell.max_row and merged_cell.min_col <= 11 <= merged_cell.max_col:
                ws.cell(row=merged_cell.min_row, column=merged_cell.min_col, value=', '.join(observaciones_dict[baja.motivo]))
                assigned = True
                break
        if not assigned:
            ws[f'L{fila}'] = ', '.join(observaciones_dict[baja.motivo])


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Reporte.xlsx'
    wb.save(response)

    return response